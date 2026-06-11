"""Derive country-level capacity weights from TYNDP 2024 PEMMDB files."""

from __future__ import annotations

import argparse
from collections import defaultdict
from pathlib import Path
import re

import pandas as pd
from openpyxl import load_workbook


REPO = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT = REPO / "outputs" / "sensitivity" / "country_weights_from_tyndp2024_pemmdb_nt2040.csv"
DEFAULT_SOURCE_URL = "https://2024-data.entsos-tyndp-scenarios.eu/files/scenarios-inputs/PEMMDB2.zip"

NODE_PREFIX_TO_COUNTRY = {
    "AT": "Austria",
    "BE": "Belgium",
    "CH": "Switzerland",
    "CZ": "Czechia",
    "DE": "Germany",
    "ES": "Spain",
    "FR": "France",
    "IT": "Italy",
    "LU": "Luxembourg",
    "NL": "Netherlands",
    "PL": "Poland",
}

WEIGHTING_ORDER = [
    "capacity",
    "renewables",
    "solar",
    "pv",
    "wind",
    "wind_onshore",
    "wind_offshore",
    "hydro",
    "pumped_hydro",
    "storage_total",
    "bio",
    "nuclear",
    "storage",
    "thermal",
]


def main(argv: list[str] | None = None) -> None:
    args = parse_args(argv)
    rows = derive_weights(args.pemmdb_root, args.year, args.scenario, args.source_url)
    result = pd.DataFrame(rows).sort_values(["weighting", "country"])
    args.output.parent.mkdir(parents=True, exist_ok=True)
    result.to_csv(args.output, index=False)
    print(args.output)


def derive_weights(pemmdb_root: Path, year: int, scenario: str, source_url: str) -> list[dict[str, object]]:
    year_dir = resolve_year_dir(pemmdb_root, year)
    pattern = f"PEMMDB_*_{scenario}_{year}.xlsx"
    by_country: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    nodes_by_country: dict[str, set[str]] = defaultdict(set)

    for workbook_path in sorted(year_dir.glob(pattern)):
        node = parse_node(workbook_path.name, scenario, year)
        country = map_node_to_country(node)
        if country is None:
            continue
        components = extract_components(workbook_path)
        nodes_by_country[country].add(node)
        for group, value in aggregate_components(components).items():
            by_country[country][group] += value

    rows = []
    for group in WEIGHTING_ORDER:
        weighting = f"{group}_tyndp2024_pemmdb_nt{year}"
        for country, values in by_country.items():
            weight = float(values.get(group, 0.0))
            if weight <= 0:
                continue
            rows.append(
                {
                    "weighting": weighting,
                    "technology_group": group,
                    "country": country,
                    "weight": weight,
                    "unit": "MW",
                    "source_year": year,
                    "scenario": scenario,
                    "source_dataset": (
                        "ENTSO-E/ENTSOG TYNDP 2024 Scenarios final package, "
                        f"PEMMDB 2.5, {format_scenario_name(scenario)} {year}"
                    ),
                    "source_url": source_url,
                    "nodes": "+".join(sorted(nodes_by_country[country])),
                }
            )
    if not rows:
        raise ValueError(f"No TYNDP PEMMDB weights found in {year_dir}")
    return rows


def resolve_year_dir(pemmdb_root: Path, year: int) -> Path:
    candidates = [
        pemmdb_root / str(year),
        pemmdb_root / "PEMMDB2" / str(year),
        pemmdb_root / "PEMMDB2" / "PEMMDB2" / str(year),
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    raise FileNotFoundError(f"Cannot find PEMMDB year directory for {year} below {pemmdb_root}")


def parse_node(filename: str, scenario: str, year: int) -> str:
    match = re.match(rf"PEMMDB_(.+?)_{re.escape(scenario)}_{year}\.xlsx$", filename)
    if not match:
        raise ValueError(f"Unexpected PEMMDB filename: {filename}")
    return match.group(1)


def map_node_to_country(node: str) -> str | None:
    return NODE_PREFIX_TO_COUNTRY.get(node[:2])


def format_scenario_name(scenario: str) -> str:
    return {"NationalTrends": "National Trends"}.get(scenario, scenario)


def extract_components(path: Path) -> dict[str, float]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    components: dict[str, float] = defaultdict(float)
    components.update(extract_thermal(workbook))
    components["other_non_res"] = extract_other_non_res(workbook)
    components.update(extract_wind(workbook))
    components.update(extract_solar(workbook))
    components.update(extract_hydro(workbook))
    other_res, bio = extract_other_res(workbook)
    components["other_res"] = other_res
    components["bio"] = bio
    components["storage"] = extract_battery(workbook)
    return components


def extract_thermal(workbook) -> dict[str, float]:
    if "Thermal" not in workbook.sheetnames:
        return {"thermal": 0.0, "nuclear": 0.0}
    sheet = workbook["Thermal"]
    rows = list(sheet.iter_rows(values_only=True))
    header_idx = find_row(rows, "Net generating capacity")
    if header_idx is None:
        return {"thermal": 0.0, "nuclear": 0.0}
    header = rows[header_idx]
    capacity_col = find_col(header, "Net generating capacity")
    fuel = None
    result = {"thermal": 0.0, "nuclear": 0.0}
    for row in rows[header_idx + 1 :]:
        if row[0] not in (None, ""):
            fuel = str(row[0]).strip()
        capacity = to_float(row[capacity_col] if capacity_col < len(row) else None)
        if capacity <= 0 or not fuel:
            continue
        if "nuclear" in fuel.lower():
            result["nuclear"] += capacity
        else:
            result["thermal"] += capacity
    return result


def extract_other_non_res(workbook) -> float:
    if "Other Non-RES" not in workbook.sheetnames:
        return 0.0
    sheet = workbook["Other Non-RES"]
    total = 0.0
    for row in sheet.iter_rows(values_only=True):
        if any("Installed capacity" in str(value) for value in row if value is not None):
            total += sum(to_float(value) for value in row[2:])
            break
    return total


def extract_wind(workbook) -> dict[str, float]:
    if "Wind" not in workbook.sheetnames:
        return {"wind_onshore": 0.0, "wind_offshore": 0.0}
    sheet = workbook["Wind"]
    result = {"wind_onshore": 0.0, "wind_offshore": 0.0}
    for row in sheet.iter_rows(values_only=True):
        label = str(row[0] or "").lower()
        value = to_float(row[1] if len(row) > 1 else None) * 1000.0
        if "onshore wind" in label:
            result["wind_onshore"] += value
        elif "offshore wind" in label:
            result["wind_offshore"] += value
    return result


def extract_solar(workbook) -> dict[str, float]:
    if "Solar" not in workbook.sheetnames:
        return {"solar_pv": 0.0, "solar_rooftop": 0.0, "solar_csp": 0.0}
    sheet = workbook["Solar"]
    result = {"solar_pv": 0.0, "solar_rooftop": 0.0, "solar_csp": 0.0}
    for row in sheet.iter_rows(values_only=True):
        label = str(row[0] or "").lower()
        value = to_float(row[1] if len(row) > 1 else None) * 1000.0
        if "photovoltaic" in label:
            result["solar_pv"] += value
        elif "rooftop" in label:
            result["solar_rooftop"] += value
        elif "installed capacities thermal solar" in label or "solar thermal with storage" in label:
            result["solar_csp"] += value
    return result


def extract_hydro(workbook) -> dict[str, float]:
    if "Hydro" not in workbook.sheetnames:
        return {"hydro": 0.0, "pumped_hydro": 0.0}
    sheet = workbook["Hydro"]
    result = {"hydro": 0.0, "pumped_hydro": 0.0}
    for row in sheet.iter_rows(values_only=True):
        label = str(row[0] or "").lower()
        if "turbining capacity" in label:
            value = max(to_float(row[1] if len(row) > 1 else None), 0.0)
            if "pump storage" in label:
                result["pumped_hydro"] += value
            else:
                result["hydro"] += value
    return result


def extract_other_res(workbook) -> tuple[float, float]:
    if "Other RES" not in workbook.sheetnames:
        return 0.0, 0.0
    rows = list(workbook["Other RES"].iter_rows(values_only=True))
    total = 0.0
    bio = 0.0
    for idx, row in enumerate(rows):
        label = str(row[0] or "")
        if label.startswith("Installed capacity excl."):
            total = to_float(row[1] if len(row) > 1 else None)
            if idx + 1 < len(rows):
                header = rows[idx]
                values = rows[idx + 1]
                for col, name in enumerate(header):
                    normalized = str(name or "").lower()
                    if "small biomass" in normalized or "waste" == normalized.strip():
                        bio += to_float(values[col] if col < len(values) else None)
            break
    return total, bio


def extract_battery(workbook) -> float:
    if "Battery" not in workbook.sheetnames:
        return 0.0
    rows = list(workbook["Battery"].iter_rows(values_only=True))
    header_idx = find_row(rows, "Net maximum capacity - generation")
    if header_idx is None:
        return 0.0
    capacity_col = find_col(rows[header_idx], "Net maximum capacity - generation")
    for row in rows[header_idx + 1 :]:
        if str(row[0] or "").strip().lower() == "battery":
            return to_float(row[capacity_col] if capacity_col < len(row) else None)
    return 0.0


def aggregate_components(components: dict[str, float]) -> dict[str, float]:
    wind = components["wind_onshore"] + components["wind_offshore"]
    pv = components["solar_pv"] + components["solar_rooftop"]
    solar = pv + components["solar_csp"]
    renewables = wind + solar + components["hydro"] + components["other_res"]
    storage_total = components["storage"] + components["pumped_hydro"]
    thermal = components["thermal"] + components["other_non_res"]
    capacity = renewables + thermal + components["nuclear"] + storage_total
    return {
        "capacity": capacity,
        "renewables": renewables,
        "solar": solar,
        "pv": pv,
        "wind": wind,
        "wind_onshore": components["wind_onshore"],
        "wind_offshore": components["wind_offshore"],
        "hydro": components["hydro"],
        "pumped_hydro": components["pumped_hydro"],
        "bio": components["bio"],
        "nuclear": components["nuclear"],
        "storage": components["storage"],
        "storage_total": storage_total,
        "thermal": thermal,
    }


def find_row(rows: list[tuple[object, ...]], needle: str) -> int | None:
    needle = needle.lower()
    for idx, row in enumerate(rows):
        if any(needle in str(value).lower().replace("\n", " ") for value in row if value is not None):
            return idx
    return None


def find_col(row: tuple[object, ...], needle: str) -> int:
    needle = needle.lower()
    for idx, value in enumerate(row):
        if needle in str(value).lower().replace("\n", " "):
            return idx
    raise ValueError(f"Cannot find column containing {needle!r}")


def to_float(value) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace("\xa0", "").replace(",", "").strip()
    if not text or text.startswith("#"):
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--pemmdb-root", type=Path, required=True)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--year", type=int, default=2040)
    parser.add_argument("--scenario", default="NationalTrends")
    parser.add_argument("--source-url", default=DEFAULT_SOURCE_URL)
    return parser.parse_args(argv)


if __name__ == "__main__":
    main()
